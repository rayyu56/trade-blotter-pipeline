"""Gold layer: write aggregated business outputs to formatted Excel files.

Produces per-dataset .xlsx files and a single daily summary workbook with
one tab per dataset — both written to data/gold/.

Formatting applied to every sheet:
- Bold, white-on-blue header row
- Numeric columns formatted with thousands separators
- All columns auto-sized to fit their widest value
"""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from trade_blotter.utils.logger import get_logger

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# Number formats keyed by column name (applied wherever the column appears)
_NOTIONAL_FMT = '#,##0.00'
_QUANTITY_FMT = '#,##0.0000'
_COUNT_FMT    = '#,##0'

_COLUMN_FORMATS: dict[str, str] = {
    "buy_notional":    _NOTIONAL_FMT,
    "sell_notional":   _NOTIONAL_FMT,
    "total_notional":  _NOTIONAL_FMT,
    "net_pnl":         _NOTIONAL_FMT,
    "buy_qty":         _QUANTITY_FMT,
    "sell_qty":        _QUANTITY_FMT,
    "net_position":    _QUANTITY_FMT,
    "avg_buy_price":   _QUANTITY_FMT,
    "avg_sell_price":  _QUANTITY_FMT,
    "price":           _QUANTITY_FMT,
    "trade_count":     _COUNT_FMT,
}

# Minimum and maximum column widths (in character units)
_MIN_COL_WIDTH = 8
_MAX_COL_WIDTH = 40


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _native(value: Any) -> Any:
    """Convert pandas/numpy scalars to Python-native types for openpyxl."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):          # numpy scalar → Python primitive
        return value.item()
    if isinstance(value, pd.Timestamp): # pandas Timestamp → Python datetime
        return value.to_pydatetime()
    return value


def _write_sheet(ws, df: pd.DataFrame) -> None:
    """Populate an openpyxl worksheet from *df* with headers and formatting."""
    columns = list(df.columns)

    # ── Header row ──────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, (col_name, raw) in enumerate(zip(columns, row), start=1):
            value = _native(raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in _COLUMN_FORMATS and value is not None:
                cell.number_format = _COLUMN_FORMATS[col_name]

    # ── Auto-size columns ────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_chars = len(col_name)
        for row_idx in range(2, len(df) + 2):
            raw = ws.cell(row=row_idx, column=col_idx).value
            if raw is not None:
                max_chars = max(max_chars, len(str(raw)))
        width = max(_MIN_COL_WIDTH, min(max_chars + 4, _MAX_COL_WIDTH))
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_excel(
    df: pd.DataFrame,
    output_dir: str | Path,
    dataset_name: str,
) -> Path:
    """Write a gold DataFrame to a single-sheet formatted Excel file.

    Args:
        df:           Gold-layer DataFrame to persist.
        output_dir:   Directory to write into (``data/gold/`` by default).
        dataset_name: File stem — e.g. ``"pnl_by_symbol"`` →
                      ``data/gold/pnl_by_symbol.xlsx``.

    Returns:
        Path to the written file.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{dataset_name}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = dataset_name

    _write_sheet(ws, df)
    wb.save(path)
    logger.info("Wrote %d rows to %s", len(df), path)
    return path


def write_daily_summary(
    outputs: dict[str, pd.DataFrame],
    output_dir: str | Path,
    date_str: str | None = None,
) -> Path:
    """Write a single workbook with one tab per gold dataset.

    Args:
        outputs:    Mapping of ``dataset_name → DataFrame`` for all five
                    gold outputs (pnl_by_symbol, pnl_by_trader, pnl_summary,
                    net_positions, position_snapshot).
        output_dir: Directory to write into (``data/gold/`` by default).
        date_str:   ``YYYY-MM-DD`` date string used in the filename.
                    Defaults to today's date.

    Returns:
        Path to the written workbook (``daily_summary_YYYY-MM-DD.xlsx``).
    """
    if date_str is None:
        date_str = datetime.date.today().isoformat()

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"daily_summary_{date_str}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)   # drop the default blank sheet

    for name, df in outputs.items():
        ws = wb.create_sheet(title=name)
        _write_sheet(ws, df)

    wb.save(path)
    logger.info("Daily summary written to %s (%d tab(s))", path, len(outputs))
    return path
