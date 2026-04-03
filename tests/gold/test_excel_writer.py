"""Tests for the gold Excel writer layer."""

import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from trade_blotter.gold.excel_writer import write_daily_summary, write_excel


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def pnl_df() -> pd.DataFrame:
    return pd.DataFrame({
        "symbol":          ["AAPL", "MSFT"],
        "asset_class":     ["Equity", "Equity"],
        "currency":        ["USD", "USD"],
        "buy_notional":    [100_000.0, 50_000.0],
        "sell_notional":   [80_000.0, 60_000.0],
        "total_notional":  [180_000.0, 110_000.0],
        "net_pnl":         [-20_000.0, 10_000.0],
        "trade_count":     [5, 3],
    })


@pytest.fixture
def positions_df() -> pd.DataFrame:
    return pd.DataFrame({
        "symbol":       ["AAPL"],
        "asset_class":  ["Equity"],
        "currency":     ["USD"],
        "buy_qty":      [100.0],
        "sell_qty":     [40.0],
        "net_position": [60.0],
        "trade_count":  [2],
        "as_of_date":   [datetime.date(2024, 1, 15)],
    })


@pytest.fixture
def all_outputs(pnl_df, positions_df) -> dict:
    return {
        "pnl_by_symbol":    pnl_df,
        "pnl_by_trader":    pnl_df[["symbol", "total_notional", "net_pnl", "trade_count"]].copy(),
        "pnl_summary":      pnl_df[["asset_class", "currency", "total_notional", "net_pnl", "trade_count"]].copy(),
        "net_positions":    positions_df.drop(columns=["as_of_date"]),
        "position_snapshot": positions_df,
    }


# ---------------------------------------------------------------------------
# write_excel — file creation and path handling
# ---------------------------------------------------------------------------

class TestWriteExcelFileCreation:
    def test_creates_xlsx_file(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        assert path.exists()

    def test_suffix_is_xlsx(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        assert path.suffix == ".xlsx"

    def test_filename_matches_dataset_name(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        assert path.name == "pnl_by_symbol.xlsx"

    def test_returns_path_object(self, tmp_path, pnl_df):
        result = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        assert isinstance(result, Path)

    def test_creates_output_dir_if_missing(self, tmp_path, pnl_df):
        new_dir = tmp_path / "gold" / "nested"
        write_excel(pnl_df, new_dir, "test")
        assert new_dir.exists()


# ---------------------------------------------------------------------------
# write_excel — header formatting
# ---------------------------------------------------------------------------

class TestWriteExcelHeaders:
    def test_header_row_is_bold(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        for cell in ws[1]:
            assert cell.font.bold

    def test_header_column_names_match_dataframe(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        headers = [ws.cell(row=1, column=i).value for i in range(1, len(pnl_df.columns) + 1)]
        assert headers == list(pnl_df.columns)

    def test_header_fill_is_applied(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        first_header = ws.cell(row=1, column=1)
        # openpyxl may return "FF1F4E79" or "001F4E79" depending on version
        assert first_header.fill.fgColor.rgb.endswith("1F4E79")

    def test_sheet_title_matches_dataset_name(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        wb = load_workbook(path)
        assert "pnl_by_symbol" in wb.sheetnames


# ---------------------------------------------------------------------------
# write_excel — data content
# ---------------------------------------------------------------------------

class TestWriteExcelData:
    def test_row_count_equals_dataframe_plus_header(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        assert ws.max_row == len(pnl_df) + 1

    def test_string_values_written_correctly(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        assert ws.cell(row=2, column=1).value == "AAPL"
        assert ws.cell(row=3, column=1).value == "MSFT"

    def test_numeric_values_written_correctly(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
        net_pnl_col = headers.index("net_pnl") + 1
        assert ws.cell(row=2, column=net_pnl_col).value == pytest.approx(-20_000.0)

    def test_date_column_written_without_error(self, tmp_path, positions_df):
        # Should not raise even when an as_of_date column is present
        path = write_excel(positions_df, tmp_path, "position_snapshot")
        assert path.exists()

    def test_nan_values_written_as_empty(self, tmp_path):
        df = pd.DataFrame({"symbol": ["AAPL"], "net_pnl": [float("nan")]})
        path = write_excel(df, tmp_path, "test")
        ws = load_workbook(path).active
        assert ws.cell(row=2, column=2).value is None


# ---------------------------------------------------------------------------
# write_excel — number formatting
# ---------------------------------------------------------------------------

class TestWriteExcelNumberFormats:
    def _col_index(self, ws, col_name: str) -> int:
        return [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)].index(col_name) + 1

    def test_notional_column_format(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        col = self._col_index(ws, "net_pnl")
        assert ws.cell(row=2, column=col).number_format == "#,##0.00"

    def test_buy_notional_format(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        col = self._col_index(ws, "buy_notional")
        assert ws.cell(row=2, column=col).number_format == "#,##0.00"

    def test_trade_count_format(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        col = self._col_index(ws, "trade_count")
        assert ws.cell(row=2, column=col).number_format == "#,##0"

    def test_quantity_column_format(self, tmp_path, positions_df):
        path = write_excel(positions_df, tmp_path, "net_positions")
        ws = load_workbook(path).active
        col = self._col_index(ws, "net_position")
        assert ws.cell(row=2, column=col).number_format == "#,##0.0000"


# ---------------------------------------------------------------------------
# write_excel — column sizing
# ---------------------------------------------------------------------------

class TestWriteExcelColumnSizing:
    def test_all_columns_have_explicit_width(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        for i in range(1, len(pnl_df.columns) + 1):
            letter = get_column_letter(i)
            assert ws.column_dimensions[letter].width > 0

    def test_width_at_least_as_wide_as_header(self, tmp_path, pnl_df):
        path = write_excel(pnl_df, tmp_path, "pnl_by_symbol")
        ws = load_workbook(path).active
        for i, col_name in enumerate(pnl_df.columns, start=1):
            letter = get_column_letter(i)
            assert ws.column_dimensions[letter].width >= len(col_name)


# ---------------------------------------------------------------------------
# write_daily_summary — file creation
# ---------------------------------------------------------------------------

class TestWriteDailySummaryFileCreation:
    def test_creates_xlsx_file(self, tmp_path, all_outputs):
        path = write_daily_summary(all_outputs, tmp_path, date_str="2024-01-15")
        assert path.exists()

    def test_suffix_is_xlsx(self, tmp_path, all_outputs):
        path = write_daily_summary(all_outputs, tmp_path, date_str="2024-01-15")
        assert path.suffix == ".xlsx"

    def test_filename_contains_date(self, tmp_path, pnl_df):
        path = write_daily_summary({"pnl_by_symbol": pnl_df}, tmp_path, date_str="2024-01-15")
        assert "2024-01-15" in path.name

    def test_default_date_is_today(self, tmp_path, pnl_df):
        path = write_daily_summary({"pnl_by_symbol": pnl_df}, tmp_path)
        assert datetime.date.today().isoformat() in path.name

    def test_returns_path_object(self, tmp_path, pnl_df):
        result = write_daily_summary({"pnl_by_symbol": pnl_df}, tmp_path, date_str="2024-01-15")
        assert isinstance(result, Path)

    def test_creates_output_dir_if_missing(self, tmp_path, pnl_df):
        new_dir = tmp_path / "gold" / "reports"
        write_daily_summary({"pnl_by_symbol": pnl_df}, new_dir, date_str="2024-01-15")
        assert new_dir.exists()


# ---------------------------------------------------------------------------
# write_daily_summary — sheet structure
# ---------------------------------------------------------------------------

class TestWriteDailySummarySheets:
    def test_all_five_tabs_present(self, tmp_path, all_outputs):
        path = write_daily_summary(all_outputs, tmp_path, date_str="2024-01-15")
        wb = load_workbook(path)
        assert set(wb.sheetnames) == set(all_outputs.keys())

    def test_tab_count_matches_outputs(self, tmp_path, pnl_df):
        outputs = {"pnl_by_symbol": pnl_df, "pnl_by_trader": pnl_df, "pnl_summary": pnl_df}
        path = write_daily_summary(outputs, tmp_path, date_str="2024-01-15")
        wb = load_workbook(path)
        assert len(wb.sheetnames) == 3

    def test_no_default_blank_sheet(self, tmp_path, pnl_df):
        path = write_daily_summary({"pnl_by_symbol": pnl_df}, tmp_path, date_str="2024-01-15")
        wb = load_workbook(path)
        assert "Sheet" not in wb.sheetnames

    def test_tab_headers_are_bold(self, tmp_path, pnl_df):
        path = write_daily_summary({"pnl_by_symbol": pnl_df}, tmp_path, date_str="2024-01-15")
        ws = load_workbook(path)["pnl_by_symbol"]
        for cell in ws[1]:
            assert cell.font.bold

    def test_tab_data_matches_dataframe(self, tmp_path, pnl_df):
        path = write_daily_summary({"pnl_by_symbol": pnl_df}, tmp_path, date_str="2024-01-15")
        ws = load_workbook(path)["pnl_by_symbol"]
        headers = [ws.cell(row=1, column=i).value for i in range(1, len(pnl_df.columns) + 1)]
        assert headers == list(pnl_df.columns)
        assert ws.max_row == len(pnl_df) + 1

    def test_position_snapshot_tab_present(self, tmp_path, positions_df):
        path = write_daily_summary({"position_snapshot": positions_df}, tmp_path, date_str="2024-01-15")
        wb = load_workbook(path)
        assert "position_snapshot" in wb.sheetnames
